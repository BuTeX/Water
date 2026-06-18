from __future__ import annotations

import argparse
import hashlib
import json
import sqlite3
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from init_db import DB_PATH, init_db


PROJECT_DIR = Path(__file__).resolve().parents[2]
DEFAULT_RAW_DIR = PROJECT_DIR / "data" / "raw"
FIRST_MONTH = "2025-05"


@dataclass(frozen=True)
class PaymentRow:
    paid_at: date
    amount: int
    house: int
    row_number: int


@dataclass(frozen=True)
class ExpenseRow:
    spent_at: date
    amount: int
    title: str
    row_number: int


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def as_amount(value: Any) -> int | None:
    if isinstance(value, (int, float)):
        return int(round(value))
    return None


def month_key(value: date) -> str:
    return f"{value.year:04d}-{value.month:02d}"


def add_month(month: str, delta: int = 1) -> str:
    year, month_number = (int(part) for part in month.split("-"))
    month_number += delta
    while month_number > 12:
        year += 1
        month_number -= 12
    while month_number < 1:
        year -= 1
        month_number += 12
    return f"{year:04d}-{month_number:02d}"


def month_range(start: str, end: str) -> list[str]:
    months: list[str] = []
    current = start
    while current <= end:
        months.append(current)
        current = add_month(current)
    return months


def access_code_for_house(number: int) -> str:
    digest = hashlib.sha256(f"water-house-{number}".encode("utf-8")).hexdigest()[:12]
    return f"h{number}-{digest}"


def base_amount(month: str) -> int:
    return 1000 if month >= "2026-07" else 500


def charge_amount(month: str, extra_by_month: dict[str, int]) -> int:
    return base_amount(month) + extra_by_month.get(month, 0)


def find_workbook(path_arg: str | None) -> Path:
    if path_arg:
        path = Path(path_arg)
        if not path.is_absolute():
            path = Path.cwd() / path
        return path

    matches = sorted(DEFAULT_RAW_DIR.glob("*.xlsx"))
    if not matches:
        raise FileNotFoundError(f"No .xlsx files found in {DEFAULT_RAW_DIR}")
    return matches[0]


def read_payments(ws) -> tuple[list[PaymentRow], list[dict[str, Any]]]:
    payments: list[PaymentRow] = []
    rejected: list[dict[str, Any]] = []

    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(value in (None, "") for value in row):
            continue

        paid_at = as_date(row[0])
        amount = as_amount(row[1])
        raw_house = row[3] if len(row) > 3 else None

        try:
            house = int(raw_house) if raw_house is not None else None
        except (TypeError, ValueError):
            house = None

        if paid_at is None or amount is None or house is None:
            rejected.append(
                {
                    "sheet": ws.title,
                    "row": index,
                    "reason": "missing payment date, amount or house number",
                    "values": [str(value) if value is not None else None for value in row[:5]],
                }
            )
            continue

        payments.append(PaymentRow(paid_at=paid_at, amount=amount, house=house, row_number=index))

    return payments, rejected


def read_expenses(ws) -> tuple[list[ExpenseRow], list[dict[str, Any]]]:
    expenses: list[ExpenseRow] = []
    rejected: list[dict[str, Any]] = []

    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(value in (None, "") for value in row):
            continue

        spent_at = as_date(row[0])
        amount = as_amount(row[1])
        title = str(row[2] or "").strip() if len(row) > 2 else ""

        if spent_at is None or amount is None or not title:
            rejected.append(
                {
                    "sheet": ws.title,
                    "row": index,
                    "reason": "missing expense date, amount or title",
                    "values": [str(value) if value is not None else None for value in row[:4]],
                }
            )
            continue

        expenses.append(ExpenseRow(spent_at=spent_at, amount=amount, title=title, row_number=index))

    return expenses, rejected


def reset_imported_data(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        DELETE FROM payment_allocations;
        DELETE FROM payments;
        DELETE FROM expenses;
        DELETE FROM houses;
        DELETE FROM sqlite_sequence WHERE name IN ('payment_allocations', 'payments', 'expenses', 'houses');
        """
    )


def insert_houses(conn: sqlite3.Connection, payments: list[PaymentRow]) -> dict[int, int]:
    first_payment_by_house: dict[int, date] = {}
    for payment in sorted(payments, key=lambda item: (item.house, item.paid_at)):
        first_payment_by_house.setdefault(payment.house, payment.paid_at)

    house_ids: dict[int, int] = {}
    for house in sorted(first_payment_by_house):
        start_month = max(month_key(first_payment_by_house[house]), FIRST_MONTH)
        conn.execute(
            """
            INSERT INTO houses (number, display_name, status, starts_on, access_code)
            VALUES (?, ?, 'active', ?, ?)
            ON CONFLICT(number) DO UPDATE SET
              display_name = excluded.display_name,
              status = 'active',
              starts_on = excluded.starts_on,
              updated_at = CURRENT_TIMESTAMP
            """,
            (house, f"ул. Уютная {house}", start_month, access_code_for_house(house)),
        )

    for row in conn.execute("SELECT id, number FROM houses"):
        house_ids[int(row["number"])] = int(row["id"])
    return house_ids


def insert_expenses(conn: sqlite3.Connection, expenses: list[ExpenseRow]) -> None:
    category_id = conn.execute("SELECT id FROM expense_categories WHERE name = 'прочее'").fetchone()["id"]
    for expense in expenses:
        conn.execute(
            """
            INSERT INTO expenses (spent_at, amount, category_id, title, description_public, source)
            VALUES (?, ?, ?, ?, ?, 'excel')
            """,
            (
                expense.spent_at.isoformat(),
                expense.amount,
                category_id,
                expense.title,
                expense.title,
            ),
        )


def insert_payments_and_allocations(
    conn: sqlite3.Connection,
    payments: list[PaymentRow],
    house_ids: dict[int, int],
    as_of_month: str,
) -> None:
    extra_by_month = {
        row["month"]: int(row["amount"])
        for row in conn.execute("SELECT month, amount FROM monthly_charges WHERE kind = 'extra'")
    }

    payments_by_house: dict[int, list[PaymentRow]] = defaultdict(list)
    for payment in payments:
        payments_by_house[payment.house].append(payment)

    for house, house_payments in payments_by_house.items():
        house_id = house_ids[house]
        house_row = conn.execute("SELECT starts_on FROM houses WHERE id = ?", (house_id,)).fetchone()
        start_month = house_row["starts_on"]
        charged_by_month = {month: charge_amount(month, extra_by_month) for month in month_range(start_month, as_of_month)}
        allocated_by_month: dict[str, int] = defaultdict(int)
        next_future_month = add_month(as_of_month)

        for payment in sorted(house_payments, key=lambda item: (item.paid_at, item.row_number)):
            cursor = conn.execute(
                """
                INSERT INTO payments (house_id, paid_at, amount, method, source)
                VALUES (?, ?, ?, 'other', 'excel')
                """,
                (house_id, payment.paid_at.isoformat(), payment.amount),
            )
            payment_id = int(cursor.lastrowid)
            remaining = payment.amount

            for month in month_range(start_month, as_of_month):
                if remaining <= 0:
                    break
                outstanding = charged_by_month[month] - allocated_by_month[month]
                if outstanding <= 0:
                    continue
                allocation = min(remaining, outstanding)
                conn.execute(
                    "INSERT INTO payment_allocations (payment_id, month, amount) VALUES (?, ?, ?)",
                    (payment_id, month, allocation),
                )
                allocated_by_month[month] += allocation
                remaining -= allocation

            while remaining > 0:
                month = next_future_month
                monthly_charge = charge_amount(month, extra_by_month)
                allocation = min(remaining, monthly_charge)
                conn.execute(
                    "INSERT INTO payment_allocations (payment_id, month, amount) VALUES (?, ?, ?)",
                    (payment_id, month, allocation),
                )
                allocated_by_month[month] += allocation
                remaining -= allocation
                if allocated_by_month[month] >= monthly_charge:
                    next_future_month = add_month(next_future_month)


def build_report(
    conn: sqlite3.Connection,
    payments: list[PaymentRow],
    expenses: list[ExpenseRow],
    rejected: list[dict[str, Any]],
    as_of_month: str,
) -> dict[str, Any]:
    house_rows = conn.execute("SELECT id, number, starts_on FROM houses ORDER BY number").fetchall()
    extra_by_month = {
        row["month"]: int(row["amount"])
        for row in conn.execute("SELECT month, amount FROM monthly_charges WHERE kind = 'extra'")
    }

    house_report = []
    total_debt = 0
    total_overpaid = 0
    for house in house_rows:
        due = sum(charge_amount(month, extra_by_month) for month in month_range(house["starts_on"], as_of_month))
        paid = conn.execute("SELECT COALESCE(SUM(amount), 0) AS total FROM payments WHERE house_id = ?", (house["id"],)).fetchone()["total"]
        debt = max(due - paid, 0)
        overpaid = max(paid - due, 0)
        total_debt += debt
        total_overpaid += overpaid
        house_report.append(
            {
                "house": house["number"],
                "starts_on": house["starts_on"],
                "paid": paid,
                "due": due,
                "debt": debt,
                "overpaid": overpaid,
            }
        )

    return {
        "as_of_month": as_of_month,
        "houses": len(house_rows),
        "payments": len(payments),
        "payments_total": sum(item.amount for item in payments),
        "expenses": len(expenses),
        "expenses_total": sum(item.amount for item in expenses),
        "balance": sum(item.amount for item in payments) - sum(item.amount for item in expenses),
        "total_debt": total_debt,
        "total_overpaid": total_overpaid,
        "rejected_rows": rejected,
        "house_report": house_report,
    }


def import_workbook(workbook_path: Path, reset: bool, as_of_month: str) -> dict[str, Any]:
    init_db()
    workbook = load_workbook(workbook_path, data_only=False)
    payments, rejected_payments = read_payments(workbook["Пополнения"])
    expenses, rejected_expenses = read_expenses(workbook["Список расходов"])
    rejected = rejected_payments + rejected_expenses

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        with conn:
            if reset:
                reset_imported_data(conn)
            house_ids = insert_houses(conn, payments)
            insert_expenses(conn, expenses)
            insert_payments_and_allocations(conn, payments, house_ids, as_of_month)
        return build_report(conn, payments, expenses, rejected, as_of_month)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", nargs="?", help="Path to source .xlsx. Defaults to first file in data/raw.")
    parser.add_argument("--reset", action="store_true", help="Delete imported houses, payments and expenses before import.")
    parser.add_argument("--as-of-month", default="2026-06", help="Last due month for import checks, YYYY-MM.")
    args = parser.parse_args()

    workbook_path = find_workbook(args.workbook)
    report = import_workbook(workbook_path, args.reset, args.as_of_month)
    report_path = DB_PATH.parent / "last-import-report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Imported workbook: {workbook_path}")
    print(f"As of month: {report['as_of_month']}")
    print(f"Houses: {report['houses']}")
    print(f"Payments: {report['payments']} / {report['payments_total']} RUB")
    print(f"Expenses: {report['expenses']} / {report['expenses_total']} RUB")
    print(f"Balance: {report['balance']} RUB")
    print(f"Debt: {report['total_debt']} RUB")
    print(f"Overpaid: {report['total_overpaid']} RUB")
    print(f"Rejected rows: {len(report['rejected_rows'])}")
    print(f"Report: {report_path}")


if __name__ == "__main__":
    main()
