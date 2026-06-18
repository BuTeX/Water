from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def as_amount(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    return None


def month_key(value: date) -> str:
    return f"{value.year:04d}-{value.month:02d}"


def usual_amount(month: str) -> int:
    return 1000 if month >= "2026-07" else 500


def rub(value: float) -> int:
    return int(round(value))


def top_counter(counter: Counter, limit: int = 8) -> list[dict[str, Any]]:
    return [{"value": key, "count": count} for key, count in counter.most_common(limit)]


def number_from_address(address: str) -> int | None:
    match = re.search(r"(\d+)", address)
    if not match:
        return None
    return int(match.group(1))


@dataclass
class Payment:
    paid_at: date
    amount: float
    house: int


@dataclass
class Expense:
    spent_at: date
    amount: float
    title: str


def read_payments(ws) -> list[Payment]:
    payments: list[Payment] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        paid_at = as_date(row[0])
        amount = as_amount(row[1])
        house = row[3]
        if paid_at is None or amount is None or house is None:
            continue
        try:
            house_number = int(house)
        except (TypeError, ValueError):
            continue
        payments.append(Payment(paid_at=paid_at, amount=amount, house=house_number))
    return payments


def read_expenses(ws) -> list[Expense]:
    expenses: list[Expense] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        spent_at = as_date(row[0])
        amount = as_amount(row[1])
        title = row[2]
        if spent_at is None or amount is None:
            continue
        expenses.append(Expense(spent_at=spent_at, amount=amount, title=str(title or "").strip()))
    return expenses


def read_summary_sheet(ws, raw_payment_houses: set[int]) -> list[dict[str, Any]]:
    month_headers: list[tuple[int, str]] = []
    for col in range(3, ws.max_column + 1):
        header = as_date(ws.cell(row=1, column=col).value)
        if header is not None:
            month_headers.append((col, month_key(header)))

    summaries: dict[str, dict[str, Any]] = {}
    for _, month in month_headers:
        summaries[month] = {
            "month": month,
            "usual_amount": usual_amount(month),
            "summary_total": 0,
            "houses_with_any_value": 0,
            "houses_above_usual": 0,
            "amount_frequencies": Counter(),
        }

    for row in range(2, 28):
        address = ws.cell(row=row, column=1).value
        if not isinstance(address, str):
            continue

        house = number_from_address(address)
        inactive_text = any(token in address.lower() for token in ["не участвует", "не продан", "не построен"])
        is_active_like = house in raw_payment_houses or not inactive_text
        if not is_active_like:
            continue

        for col, month in month_headers:
            amount = as_amount(ws.cell(row=row, column=col).value)
            if amount is None or amount == 0:
                continue
            summary = summaries[month]
            summary["summary_total"] += amount
            summary["houses_with_any_value"] += 1
            if amount > summary["usual_amount"]:
                summary["houses_above_usual"] += 1
            summary["amount_frequencies"][rub(amount)] += 1

    normalized = []
    for summary in summaries.values():
        normalized.append(
            {
                **{key: value for key, value in summary.items() if key != "amount_frequencies"},
                "summary_total": rub(summary["summary_total"]),
                "amount_frequencies": top_counter(summary["amount_frequencies"]),
            }
        )
    return normalized


def analyze(workbook_path: Path) -> dict[str, Any]:
    wb = load_workbook(workbook_path, data_only=False)
    sheet_names = wb.sheetnames
    payments = read_payments(wb["Пополнения"])
    expenses = read_expenses(wb["Список расходов"])
    raw_payment_houses = {payment.house for payment in payments}
    summary_rows = read_summary_sheet(wb["Таблица оплат"], raw_payment_houses)

    payments_by_month: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "total_paid": 0.0,
            "payment_count": 0,
            "houses": set(),
            "payment_amount_frequencies": Counter(),
        }
    )
    house_month_totals: dict[tuple[int, str], float] = defaultdict(float)
    for payment in payments:
        month = month_key(payment.paid_at)
        row = payments_by_month[month]
        row["total_paid"] += payment.amount
        row["payment_count"] += 1
        row["houses"].add(payment.house)
        row["payment_amount_frequencies"][rub(payment.amount)] += 1
        house_month_totals[(payment.house, month)] += payment.amount

    expenses_by_month: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"total_expenses": 0.0, "expense_count": 0, "titles": []}
    )
    for expense in expenses:
        month = month_key(expense.spent_at)
        row = expenses_by_month[month]
        row["total_expenses"] += expense.amount
        row["expense_count"] += 1
        row["titles"].append({"date": expense.spent_at.isoformat(), "amount": rub(expense.amount), "title": expense.title})

    per_house_month_amounts: dict[str, list[float]] = defaultdict(list)
    high_house_months: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for (house, month), amount in sorted(house_month_totals.items()):
        per_house_month_amounts[month].append(amount)
        if amount > usual_amount(month):
            high_house_months[month].append({"house": house, "amount": rub(amount)})

    months = sorted(set(payments_by_month) | set(expenses_by_month) | {row["month"] for row in summary_rows})
    monthly = []
    for month in months:
        payment = payments_by_month.get(month)
        expense = expenses_by_month.get(month)
        amounts = per_house_month_amounts.get(month, [])
        high_houses = high_house_months.get(month, [])
        amount_counter = Counter(rub(amount) for amount in amounts)
        unique_houses_paid = len(payment["houses"]) if payment else 0
        total_paid = payment["total_paid"] if payment else 0.0
        total_expenses = expense["total_expenses"] if expense else 0.0
        usual = usual_amount(month)
        likely_extra_collection = bool(
            total_expenses > 0
            and len(high_houses) >= 3
            and amounts
            and (sum(amounts) / len(amounts)) > usual * 1.4
        )
        monthly.append(
            {
                "month": month,
                "usual_amount": usual,
                "total_paid_raw": rub(total_paid),
                "payment_count": payment["payment_count"] if payment else 0,
                "unique_houses_paid": unique_houses_paid,
                "house_month_amount_frequencies": top_counter(amount_counter),
                "houses_above_usual_count": len(high_houses),
                "houses_above_usual": sorted(high_houses, key=lambda item: (-item["amount"], item["house"]))[:12],
                "total_expenses": rub(total_expenses),
                "expense_count": expense["expense_count"] if expense else 0,
                "expense_titles": expense["titles"] if expense else [],
                "likely_extra_collection_month": likely_extra_collection,
            }
        )

    highlighted_summary_months = [
        row
        for row in summary_rows
        if row["summary_total"] > 0 and (row["houses_above_usual"] >= 3 or row["summary_total"] >= row["usual_amount"] * 10)
    ]

    return {
        "workbook": str(workbook_path),
        "sheet_names": sheet_names,
        "analyzed_sheets": sheet_names[:3],
        "skipped_sheets": sheet_names[3:],
        "payments": {
            "row_count": len(payments),
            "date_min": min(payment.paid_at for payment in payments).isoformat() if payments else None,
            "date_max": max(payment.paid_at for payment in payments).isoformat() if payments else None,
            "unique_houses": sorted(raw_payment_houses),
            "unique_houses_count": len(raw_payment_houses),
            "total": rub(sum(payment.amount for payment in payments)),
        },
        "expenses": {
            "row_count": len(expenses),
            "date_min": min(expense.spent_at for expense in expenses).isoformat() if expenses else None,
            "date_max": max(expense.spent_at for expense in expenses).isoformat() if expenses else None,
            "total": rub(sum(expense.amount for expense in expenses)),
        },
        "monthly_analysis": monthly,
        "summary_sheet_highlighted_months": highlighted_summary_months,
    }


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Usage: analyze_water_excel.py <workbook.xlsx>")

    result = analyze(Path(sys.argv[1]))
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
