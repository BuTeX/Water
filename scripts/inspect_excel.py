from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def row_values(row: tuple[Any, ...]) -> list[Any]:
    return [clean_value(cell.value) for cell in row]


def compact_row(values: list[Any]) -> list[Any]:
    last = 0
    for index, value in enumerate(values):
        if value not in (None, ""):
            last = index + 1
    return values[:last]


def inspect_sheet(ws, max_rows: int = 80, max_cols: int = 30) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    non_empty_rows = 0

    for row_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows), max_col=min(ws.max_column, max_cols)),
        start=1,
    ):
        values = compact_row(row_values(row))
        if any(value not in (None, "") for value in values):
            non_empty_rows += 1
            rows.append({"row": row_index, "values": values})

    return {
        "title": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_ranges": [str(item) for item in ws.merged_cells.ranges],
        "non_empty_rows_in_preview": non_empty_rows,
        "preview_rows": rows,
    }


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Usage: inspect_excel.py <workbook.xlsx>")

    workbook_path = Path(sys.argv[1])
    wb = load_workbook(workbook_path, data_only=False)
    sheet_names = wb.sheetnames
    target_sheet_names = sheet_names[:3]

    result = {
        "workbook": str(workbook_path),
        "all_sheet_names": sheet_names,
        "analyzed_sheet_names": target_sheet_names,
        "skipped_sheet_names": sheet_names[3:],
        "sheets": [inspect_sheet(wb[name]) for name in target_sheet_names],
    }

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
